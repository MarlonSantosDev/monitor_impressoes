"""
Monitor de filas de impressão Windows.

Registra cada job em um Excel do dia na raiz (pasta onde o .exe roda):
  log_impressoes_DDMMYYYY.xlsx  (ex.: log_impressoes_20022026.xlsx)
Cria a pasta arquivos/ e, se SALVAR_COPIA_SPL = True, grava lá uma cópia do
arquivo de spool (.SPL) do job — não é o documento original (PDF/DOC), e sim
o formato que o Windows envia para a impressora. Pode exigir permissões de admin.
Se CONVERTER_EMF_PARA_IMAGEM = True, renderiza cada página do SPL (formato EMF)
para um arquivo BMP abrível diretamente no Windows — sem dependências extras.
Requer Windows (pywin32) e openpyxl.
Erros e falhas de execução são registrados em erro.log na mesma pasta do .exe.
Se não coletar dados nos testes, defina DEBUG = True para ver impressoras e jobs na fila.
"""
import ctypes
import logging
import os
import re
import shutil
import struct
import sys
import threading
import time
import traceback
from datetime import datetime, timedelta
from typing import Optional

import win32con
import win32file
import win32print
import win32timezone  # Usado por pywin32 ao acessar job['Submitted']; necessário no .exe
from openpyxl import Workbook, load_workbook

# --- CONFIGURAÇÃO ---
# Base: pasta onde o .exe é executado (ou do script)
if getattr(sys, "frozen", False):
    PASTA_SCRIPT = os.path.dirname(sys.executable)
else:
    PASTA_SCRIPT = os.path.dirname(os.path.abspath(__file__))
# Saída: Excel do dia na raiz (log_impressoes_DDMMYYYY.xlsx) + pasta arquivos/
PASTA_ARQUIVOS = os.path.join(PASTA_SCRIPT, "arquivos")
ARQUIVO_LOG_ERRO = os.path.join(PASTA_SCRIPT, "erro.log")
NOME_ABA = "Impressões"
CABECALHO = [
    "ID_Job", "Usuario", "Data_Hora", "Arquivo", "Paginas", "Impressora", "Tamanho_Bytes", "Local_Arquivo"
]
DIAS_RETENCAO = 2       # Remover log_impressoes_*.xlsx com mais de 2 dias
INTERVALO_SEGUNDOS = 0.1  # Varredura das filas (mínimo prático para capturar todos os jobs)
CACHE_JOB_HORAS = 24    # Limpar do cache jobs processados há mais de 24 h
DEBUG = False           # True: mostra quantas impressoras/jobs por ciclo
SALVAR_COPIA_SPL = True  # Copiar arquivo de spool do job para pasta arquivos/ (pode exigir admin)
SPOOL_COPIES_MAX_IDADE_SEGUNDOS = 300  # Descartar temp SPL/SHD sem job correspondente após 5 min
CONVERTER_EMF_PARA_IMAGEM = True  # Renderizar SPL EMF para BMP ao lado do .spl (requer admin)
EMF_DPI = 150           # DPI das imagens BMP geradas

# Definido na inicialização: True se o processo tem acesso à pasta System32\spool\PRINTERS
_spool_acessivel: Optional[bool] = None


def _verificar_acesso_spool() -> bool:
    """
    Verifica se o processo tem acesso à pasta do spool (leitura).
    Retorna True se conseguir abrir/listar; False caso contrário.
    Evita tentativas e avisos repetidos quando não há permissão de admin.
    """
    spool_dir = os.path.join(
        os.environ.get("SystemRoot", "C:\\Windows"), "System32", "spool", "PRINTERS"
    )
    try:
        h = win32file.CreateFile(
            spool_dir,
            win32con.GENERIC_READ,
            win32con.FILE_SHARE_READ,
            None,
            win32con.OPEN_EXISTING,
            win32con.FILE_FLAG_BACKUP_SEMANTICS,
            None,
        )
        win32file.CloseHandle(h)
        return True
    except OSError:
        return False


def configurar_log_erro():
    """Configura o módulo logging para gravar em erro.log (append) com data/hora e ponto do código."""
    log = logging.getLogger("monitor_impressoes")
    log.setLevel(logging.DEBUG)
    if log.handlers:
        return log
    try:
        fh = logging.FileHandler(ARQUIVO_LOG_ERRO, mode="a", encoding="utf-8")
    except OSError:
        return log
    fh.setLevel(logging.DEBUG)
    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    fh.setFormatter(fmt)
    log.addHandler(fh)
    return log


def log_erro(ponto: str, mensagem: str, exc: Optional[BaseException] = None):
    """Registra em erro.log: ponto do código, mensagem e, se houver, exceção com traceback."""
    log = configurar_log_erro()
    texto = f"[{ponto}] {mensagem}"
    if exc is not None:
        log.error("%s | %s: %s\n%s", texto, type(exc).__name__, exc, traceback.format_exc())
    else:
        log.error(texto)


def log_aviso(ponto: str, mensagem: str):
    """Registra aviso em erro.log (falha de execução sem exceção)."""
    configurar_log_erro().warning("[%s] %s", ponto, mensagem)


def _bytes_para_kb_mb_gb(size_bytes: int) -> str:
    """Converte tamanho em bytes para string legível em B, KB, MB ou GB."""
    try:
        n = int(size_bytes)
    except (TypeError, ValueError):
        return "0 B"
    if n < 0:
        return "0 B"
    if n < 1024:
        return f"{n} B"
    if n < 1024 * 1024:
        return f"{n / 1024:.2f} KB"
    if n < 1024**3:
        return f"{n / (1024**2):.2f} MB"
    return f"{n / (1024**3):.2f} GB"


def _sanitizar_nome_arquivo(nome: str, max_len: int = 80) -> str:
    """Remove caracteres inválidos para nome de arquivo Windows."""
    invalidos = r'<>:"/\\|?*'
    for c in invalidos:
        nome = nome.replace(c, "_")
    nome = nome.strip(". ") or "documento"
    return nome[:max_len] if len(nome) > max_len else nome


def caminho_arquivo_copia(job_id: int, nome_documento: str, impressora: str) -> str:
    """
    Retorna o caminho completo (local do arquivo) onde a cópia do spool será salva em arquivos/.
    Esse valor deve ser usado na coluna Local_Arquivo do Excel.
    """
    base = _sanitizar_nome_arquivo(nome_documento)
    impressora_safe = _sanitizar_nome_arquivo(impressora, 40)
    dest_nome = f"{job_id}_{impressora_safe}_{base}.spl"
    return os.path.join(PASTA_ARQUIVOS, dest_nome)


def copiar_spool_para_arquivos(job_id: int, dest_path: str, max_tentativas: int = 5) -> bool:
    """
    Tenta copiar o arquivo de spool do job (System32\\spool\\PRINTERS) para dest_path.
    Faz até max_tentativas com pequenas pausas (o .SPL pode aparecer um pouco depois do job).
    Retorna True se copiou, False se falhou (ex.: sem permissão, arquivo já removido).
    O arquivo é .SPL (formato spool Windows), não o documento original.
    """
    spool_dir = os.path.join(os.environ.get("SystemRoot", "C:\\Windows"), "System32", "spool", "PRINTERS")
    nome_spl = f"{job_id:05d}.SPL"
    src = os.path.join(spool_dir, nome_spl)
    for _ in range(max_tentativas):
        if os.path.isfile(src):
            try:
                shutil.copy2(src, dest_path)
                return True
            except (OSError, PermissionError):
                pass
        time.sleep(0.1)
    log_aviso(
        "copiar_spool_para_arquivos",
        f"Falha após {max_tentativas} tentativas: job_id={job_id} src={src} dest={dest_path}",
    )
    return False


def copiar_shd_para_arquivos(job_id: int, dest_spl_path: str) -> bool:
    """
    Tenta copiar o arquivo de shadow (.SHD) do job junto com o .SPL.
    O .SHD contém metadados binários completos do job (usuário, documento, páginas, tipo de dados).
    Retorna True se copiou, False se não encontrou ou falhou.
    """
    spool_dir = os.path.join(os.environ.get("SystemRoot", "C:\\Windows"), "System32", "spool", "PRINTERS")
    src = os.path.join(spool_dir, f"{job_id:05d}.SHD")
    dest = os.path.splitext(dest_spl_path)[0] + ".shd"
    if os.path.isfile(src):
        try:
            shutil.copy2(src, dest)
            return True
        except (OSError, PermissionError):
            pass
    return False


def caminho_log_do_dia(data=None):
    """Retorna o caminho do Excel do dia na raiz: log_impressoes_DDMMYYYY.xlsx"""
    if data is None:
        data = datetime.now()
    if isinstance(data, datetime):
        data_str = data.strftime("%d%m%Y")  # ex.: 20022026
    else:
        data_str = datetime.strptime(str(data), "%Y-%m-%d").strftime("%d%m%Y")
    nome = f"log_impressoes_{data_str}.xlsx"
    return os.path.join(PASTA_SCRIPT, nome)


def iniciar_log():
    """Cria a pasta arquivos/ na raiz e executa limpeza de logs antigos."""
    global _spool_acessivel
    _spool_acessivel = _verificar_acesso_spool()
    os.makedirs(PASTA_ARQUIVOS, exist_ok=True)
    print(f"Log do dia na raiz: log_impressoes_DDMMYYYY.xlsx")
    print(f"Pasta criada: arquivos/")
    if SALVAR_COPIA_SPL and not _spool_acessivel:
        print("[Aviso] Cópia do spool desativada (sem permissão em System32\\spool\\PRINTERS). Execute como Administrador para salvar cópias.")
    limpar_logs_antigos()


def limpar_logs_antigos():
    """Remove arquivos log_impressoes_DDMMYYYY.xlsx da raiz com mais de DIAS_RETENCAO dias."""
    limite = datetime.now() - timedelta(days=DIAS_RETENCAO)
    padrao = re.compile(r"^log_impressoes_(\d{8})\.xlsx$")  # DDMMYYYY
    try:
        for nome in os.listdir(PASTA_SCRIPT):
            m = padrao.match(nome)
            if not m:
                continue
            try:
                d = datetime.strptime(m.group(1), "%d%m%Y")
                if d < limite:
                    path = os.path.join(PASTA_SCRIPT, nome)
                    os.remove(path)
                    print(f"Removido log antigo (>{DIAS_RETENCAO} dias): {nome}")
            except (ValueError, OSError):
                pass
    except OSError:
        pass


# ---------------------------------------------------------------------------
# Conversão EMF → BMP
# ---------------------------------------------------------------------------

def _checar_tipo_spool(spl_path: str) -> str:
    """
    Lê os primeiros 8 bytes do SPL para detectar o tipo de dados.
    Retorna 'EMF' se for EMFSPOOL (magic=0x00000000, version=0x00010000),
    senão 'DESCONHECIDO'.
    """
    try:
        with open(spl_path, "rb") as f:
            header = f.read(8)
        if len(header) < 8:
            return "DESCONHECIDO"
        magic, version = struct.unpack_from("<II", header, 0)
        if magic == 0x00000000 and version == 0x00010000:
            return "EMF"
        return "DESCONHECIDO"
    except OSError:
        return "DESCONHECIDO"


def _extrair_emf_pages_de_spool(spl_path: str) -> list:
    """
    Parseia o container EMFSPOOL e retorna uma lista de bytes objects,
    cada um sendo um EMF stream completo para uma página.

    Estrutura EMFSPOOL:
      Header 20 bytes: ulID(4), version(4), nRecords(4), dpszDocName(4), dpszPort(4)
      Seguido de registros: iType(4), cj(4), dados[cj-8]
      iType==1 (EMRI_METAFILE) ou iType==12 (EMRI_METAFILE_DATA) = página EMF;
      blob deve iniciar com EMR_HEADER (iType==1 no 1º DWORD do blob).
      iType==2 é EMRI_PS_JOB_DATA (PostScript) — não contém EMF.
    """
    pages = []
    try:
        with open(spl_path, "rb") as f:
            data = f.read()
    except OSError:
        return pages

    if len(data) < 20:
        return pages

    magic, version, n_records, _, _ = struct.unpack_from("<IIIII", data, 0)
    if magic != 0x00000000 or version != 0x00010000:
        return pages

    pos = 20
    for _ in range(n_records):
        if pos + 8 > len(data):
            break
        iType, cj = struct.unpack_from("<II", data, pos)
        if cj < 8 or pos + cj > len(data):
            break
        if iType in (1, 12):  # EMRI_METAFILE ou EMRI_METAFILE_DATA — página EMF
            emf_blob = data[pos + 8: pos + cj]
            if len(emf_blob) >= 8:
                emr_type, _ = struct.unpack_from("<II", emf_blob, 0)
                if emr_type == 1:  # EMR_HEADER — blob válido
                    pages.append(bytes(emf_blob))
        pos += cj
        # alinhamento DWORD
        rem = pos % 4
        if rem:
            pos += 4 - rem

    return pages


def _emf_blob_para_bmp(emf_blob: bytes, dpi: int, dest_bmp_path: str) -> bool:
    """
    Renderiza um EMF blob para um arquivo BMP 24-bit usando GDI32 via ctypes.
    Não requer Pillow. Retorna True se bem-sucedido.

    Pipeline:
      SetEnhMetaFileBits → HEMF
      GetEnhMetaFileHeader → dimensões da página em 0.01mm → pixels
      GetDC(0) → screen DC de referência
      CreateCompatibleDC + CreateCompatibleBitmap → off-screen
      FillRect branco + PlayEnhMetaFile → renderiza
      GetDIBits → buffer BGR 24bpp
      Cleanup GDI
      Escrever BMP 24-bit via struct + open("wb")
    """
    gdi32 = ctypes.windll.gdi32
    user32 = ctypes.windll.user32
    DIB_RGB_COLORS = 0

    # Restype explícito: sem isso, em Windows 64-bit os handles GDI são
    # truncados de 64 para 32 bits, corrompendo todos os ponteiros.
    gdi32.SetEnhMetaFileBits.restype = ctypes.c_void_p
    gdi32.GetEnhMetaFileHeader.restype = ctypes.c_uint
    gdi32.DeleteEnhMetaFile.restype = ctypes.c_int
    gdi32.CreateCompatibleDC.restype = ctypes.c_void_p
    gdi32.CreateCompatibleBitmap.restype = ctypes.c_void_p
    gdi32.SelectObject.restype = ctypes.c_void_p
    gdi32.CreateSolidBrush.restype = ctypes.c_void_p
    gdi32.DeleteObject.restype = ctypes.c_int
    gdi32.DeleteDC.restype = ctypes.c_int
    gdi32.PlayEnhMetaFile.restype = ctypes.c_int
    gdi32.GetDIBits.restype = ctypes.c_int
    user32.GetDC.restype = ctypes.c_void_p
    user32.ReleaseDC.restype = ctypes.c_int
    user32.FillRect.restype = ctypes.c_int

    class RECT(ctypes.Structure):
        _fields_ = [("left", ctypes.c_long), ("top", ctypes.c_long),
                    ("right", ctypes.c_long), ("bottom", ctypes.c_long)]

    class BITMAPINFOHEADER(ctypes.Structure):
        _fields_ = [
            ("biSize", ctypes.c_uint32),
            ("biWidth", ctypes.c_int32),
            ("biHeight", ctypes.c_int32),
            ("biPlanes", ctypes.c_uint16),
            ("biBitCount", ctypes.c_uint16),
            ("biCompression", ctypes.c_uint32),
            ("biSizeImage", ctypes.c_uint32),
            ("biXPelsPerMeter", ctypes.c_int32),
            ("biYPelsPerMeter", ctypes.c_int32),
            ("biClrUsed", ctypes.c_uint32),
            ("biClrImportant", ctypes.c_uint32),
        ]

    try:
        hemf = gdi32.SetEnhMetaFileBits(len(emf_blob), emf_blob)
        if not hemf:
            return False
        try:
            ENHMETAHEADER_SIZE = 88
            hdr_buf = ctypes.create_string_buffer(ENHMETAHEADER_SIZE)
            if gdi32.GetEnhMetaFileHeader(hemf, ENHMETAHEADER_SIZE, hdr_buf) < ENHMETAHEADER_SIZE:
                return False

            # rclFrame: bytes 24-39 em 0.01mm (left, top, right, bottom)
            fl, ft, fr, fb = struct.unpack_from("<iiii", hdr_buf.raw, 24)
            w_01mm = fr - fl
            h_01mm = fb - ft
            if w_01mm <= 0 or h_01mm <= 0:
                return False

            # Converter 0.01mm → pixels: 1 polegada = 25.4mm = 2540 unidades de 0.01mm
            # pixels = (valor / 2540) * DPI
            width_px = max(1, int(w_01mm / 2540.0 * dpi))
            height_px = max(1, int(h_01mm / 2540.0 * dpi))

            hdc_screen = user32.GetDC(0)
            if not hdc_screen:
                return False
            try:
                hdc_mem = gdi32.CreateCompatibleDC(hdc_screen)
                hbmp = gdi32.CreateCompatibleBitmap(hdc_screen, width_px, height_px)
                hbmp_old = gdi32.SelectObject(hdc_mem, hbmp)

                # Fundo branco
                rect = RECT(0, 0, width_px, height_px)
                hbrush = gdi32.CreateSolidBrush(0x00FFFFFF)
                user32.FillRect(hdc_mem, ctypes.byref(rect), hbrush)
                gdi32.DeleteObject(hbrush)

                # Renderizar EMF
                gdi32.PlayEnhMetaFile(hdc_mem, hemf, ctypes.byref(rect))

                # Ler pixels — bottom-up (biHeight positivo = padrão BMP)
                row_stride = (width_px * 3 + 3) & ~3  # alinhamento DWORD
                bih = BITMAPINFOHEADER()
                bih.biSize = ctypes.sizeof(BITMAPINFOHEADER)
                bih.biWidth = width_px
                bih.biHeight = height_px   # positivo = bottom-up (padrão BMP)
                bih.biPlanes = 1
                bih.biBitCount = 24
                bih.biCompression = 0      # BI_RGB
                bih.biSizeImage = row_stride * height_px
                bih.biXPelsPerMeter = int(dpi / 0.0254)
                bih.biYPelsPerMeter = int(dpi / 0.0254)

                pixel_buf = ctypes.create_string_buffer(row_stride * height_px)

                # Cleanup GDI — deselecionar hbmp ANTES de GetDIBits (requisito MSDN:
                # "the bitmap must not be selected into a device context")
                gdi32.SelectObject(hdc_mem, hbmp_old)

                lines = gdi32.GetDIBits(
                    hdc_mem, hbmp, 0, height_px,
                    pixel_buf, ctypes.byref(bih), DIB_RGB_COLORS
                )

                gdi32.DeleteObject(hbmp)
                gdi32.DeleteDC(hdc_mem)

                if lines == 0:
                    return False

                # Escrever BMP sem Pillow
                dib_size = ctypes.sizeof(BITMAPINFOHEADER)
                pixel_offset = 14 + dib_size
                file_size = pixel_offset + row_stride * height_px
                bmp_file_header = struct.pack("<2sIHHI", b"BM", file_size, 0, 0, pixel_offset)

                with open(dest_bmp_path, "wb") as f:
                    f.write(bmp_file_header)
                    f.write(bytes(bih))
                    f.write(pixel_buf.raw)

                return True

            finally:
                user32.ReleaseDC(0, hdc_screen)
        finally:
            gdi32.DeleteEnhMetaFile(hemf)

    except Exception as e:
        log_erro("_emf_blob_para_bmp", f"Falha ao renderizar EMF para BMP: {dest_bmp_path}", e)
        return False


def converter_spl_para_imagens(spl_path: str, dpi: int = EMF_DPI) -> list:
    """
    Tenta converter o arquivo SPL em imagens BMP (uma por página EMF).
    Retorna lista de caminhos dos BMPs gerados (vazia se falhar ou não for EMF).
    Não lança exceção; falhas são silenciosas com log de aviso.
    Jobs em formato RAW (PostScript, PCL) são silenciosamente ignorados.
    """
    if _checar_tipo_spool(spl_path) != "EMF":
        return []
    pages = _extrair_emf_pages_de_spool(spl_path)
    if not pages:
        log_aviso("converter_spl_para_imagens", f"Nenhuma página EMF extraída de: {spl_path}")
        return []
    base = os.path.splitext(spl_path)[0]
    resultados = []
    for i, emf_blob in enumerate(pages, start=1):
        dest_bmp = f"{base}_p{i:03d}.bmp"
        if _emf_blob_para_bmp(emf_blob, dpi, dest_bmp):
            resultados.append(dest_bmp)
            if DEBUG:
                print(f"[EMF] Página {i} renderizada: {dest_bmp}")
        else:
            log_aviso("converter_spl_para_imagens", f"Falha ao renderizar página {i} de: {spl_path}")
    return resultados


# ---------------------------------------------------------------------------
# Watcher de spool em thread daemon
# ---------------------------------------------------------------------------

def watch_spool_directory(spool_copies: dict):
    """
    Thread daemon: monitora C:\\Windows\\System32\\spool\\PRINTERS\\ via ReadDirectoryChangesW.
    Ao detectar a criação de um .SPL, copia imediatamente para arquivos/_temp_{job_id}.spl
    e registra em spool_copies[job_id] = (dest_path, timestamp).
    Também copia o .SHD companion quando disponível.
    Requer permissões de administrador.
    Nota: o .SPL é formato Windows EMF/RAW, não o documento original (PDF/DOCX).
    """
    spool_dir = os.path.join(
        os.environ.get("SystemRoot", "C:\\Windows"), "System32", "spool", "PRINTERS"
    )
    try:
        h_dir = win32file.CreateFile(
            spool_dir,
            win32con.GENERIC_READ,
            win32con.FILE_SHARE_READ | win32con.FILE_SHARE_WRITE | win32con.FILE_SHARE_DELETE,
            None,
            win32con.OPEN_EXISTING,
            win32con.FILE_FLAG_BACKUP_SEMANTICS,
            None,
        )
    except OSError as e:
        log_erro("watch_spool_directory.inicio", f"Não foi possível abrir pasta do spool: {spool_dir}", e)
        print(f"[Aviso] Watcher de spool não iniciado: {e}")
        return

    print(f"[Spool] Assistindo: {spool_dir}")
    while True:
        try:
            results = win32file.ReadDirectoryChangesW(
                h_dir,
                65536,
                False,  # não recursivo
                win32con.FILE_NOTIFY_CHANGE_FILE_NAME,
                None,
                None,
            )
            for action, filename in results:
                if action != 1:  # 1 = FILE_ACTION_ADDED
                    continue
                if not filename.upper().endswith(".SPL"):
                    continue
                src = os.path.join(spool_dir, filename)
                try:
                    job_id = int(os.path.splitext(filename)[0])
                except ValueError:
                    continue
                os.makedirs(PASTA_ARQUIVOS, exist_ok=True)
                dest = os.path.join(PASTA_ARQUIVOS, f"_temp_{job_id}.spl")
                copiou = False
                for attempt in range(5):
                    try:
                        shutil.copy2(src, dest)
                        spool_copies[job_id] = (dest, time.time())
                        copiou = True
                        if DEBUG:
                            print(f"[Spool] Capturado: {filename} → {dest}")
                        break
                    except (OSError, PermissionError):
                        if attempt < 4:
                            time.sleep(0.08)
                if copiou:
                    # Copiar companion .SHD
                    src_shd = os.path.join(spool_dir, f"{job_id:05d}.SHD")
                    dest_shd = os.path.join(PASTA_ARQUIVOS, f"_temp_{job_id}.shd")
                    if os.path.isfile(src_shd):
                        try:
                            shutil.copy2(src_shd, dest_shd)
                        except (OSError, PermissionError):
                            pass
                else:
                    log_aviso("watch_spool_directory.copia", f"Cópia do spool falhou após 5 tentativas: job_id={job_id} arquivo={filename}")
                    if DEBUG:
                        print(f"[Spool] Não foi possível copiar {filename} (arquivo já removido ou bloqueado)")
        except Exception as e:
            log_erro("watch_spool_directory.loop", "Erro ao processar alterações do spool", e)
            if DEBUG:
                print(f"[Spool] Erro no watcher: {e}")
            time.sleep(1)


# ---------------------------------------------------------------------------
# Loop principal de monitoramento
# ---------------------------------------------------------------------------

def monitorar_impressoes():
    global _spool_acessivel
    if _spool_acessivel is None:
        _spool_acessivel = _verificar_acesso_spool()
    print(f"Monitorando impressões... (Pressione Ctrl+C para parar)")

    jobs_processados = {}
    excel_pendente: list = []  # [(linha_dados, job_unique_key, timestamp)]
    ultima_limpeza = datetime.now()

    # Dict compartilhado: job_id → (caminho_temp_spl, timestamp)
    spool_copies: dict = {}
    if SALVAR_COPIA_SPL and _spool_acessivel:
        t = threading.Thread(target=watch_spool_directory, args=(spool_copies,), daemon=True)
        t.start()

    # Local + conexões de rede
    flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
    primeiro_ciclo = True

    while True:
        try:
            # --- DRAIN: tentar gravar linhas pendentes de ciclos anteriores ---
            if excel_pendente:
                ainda_pendente = []
                for (linha_p, chave_p, ts_p) in excel_pendente:
                    arquivo_p = caminho_log_do_dia()
                    try:
                        if not os.path.exists(arquivo_p):
                            wb_p = Workbook()
                            ws_p = wb_p.active
                            ws_p.title = NOME_ABA
                            ws_p.append(CABECALHO)
                        else:
                            wb_p = load_workbook(arquivo_p)
                            ws_p = wb_p[NOME_ABA] if NOME_ABA in wb_p.sheetnames else wb_p.create_sheet(NOME_ABA)
                        ws_p.append(linha_p)
                        wb_p.save(arquivo_p)
                        jobs_processados[chave_p] = ts_p
                        print(f"[Pendente gravado] {chave_p}")
                    except PermissionError:
                        ainda_pendente.append((linha_p, chave_p, ts_p))
                    except Exception as e_drain:
                        log_erro("drain_pendente", f"Falha ao gravar linha pendente: {chave_p}", e_drain)
                        ainda_pendente.append((linha_p, chave_p, ts_p))
                excel_pendente = ainda_pendente

            printers = win32print.EnumPrinters(flags)
            if DEBUG or primeiro_ciclo:
                print(f"[Diagnóstico] Impressoras encontradas: {len(printers)}")
                for p in printers:
                    print(f"  - {p[2]}")
                primeiro_ciclo = False

            total_jobs_ciclo = 0
            for printer in printers:
                printer_name = printer[2]
                p_handle = None

                try:
                    p_handle = win32print.OpenPrinter(printer_name)

                    # Nível 2 traz detalhes do dono e nome do documento
                    jobs = win32print.EnumJobs(p_handle, 0, -1, 2)
                    total_jobs_ciclo += len(jobs)
                    if DEBUG and jobs:
                        print(f"[Diagnóstico] Impressora '{printer_name}': {len(jobs)} job(s) na fila")

                    for job in jobs:
                        job_id = job['JobId']
                        # Chave única: Nome da Impressora + ID do Job
                        job_unique_key = f"{printer_name}_{job_id}"

                        if job_unique_key in jobs_processados:
                            continue

                        # --- EXTRAÇÃO DE DADOS ---
                        usuario = job.get('pUserName', 'Sistema/Desconhecido')
                        documento = (
                            job.get('pDocument') or job.get('Document') or job.get('pDocName')
                            or ''
                        )
                        if isinstance(documento, str):
                            documento = documento.strip()
                        if not documento:
                            documento = f"Sem Nome (job {job_id})"
                        else:
                            documento = str(documento)
                        paginas = job.get('TotalPages', 0)
                        tamanho = job.get('Size', 0)

                        try:
                            data_raw = job.get('Submitted')
                            if data_raw is None:
                                raise ValueError("Submitted ausente")
                            data_hora = f"{data_raw.year}-{data_raw.month:02d}-{data_raw.day:02d} {data_raw.hour:02d}:{data_raw.minute:02d}:{data_raw.second:02d}"
                        except (AttributeError, TypeError, ValueError):
                            data_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                        # Local do arquivo: watcher já copiou o SPL; fallback para cópia direta
                        local_arquivo = ""
                        if SALVAR_COPIA_SPL and _spool_acessivel:
                            os.makedirs(PASTA_ARQUIVOS, exist_ok=True)
                            dest_path = caminho_arquivo_copia(job_id, documento, printer_name)
                            # Aguardar até 0,5s para o watcher capturar o SPL (thread assíncrona)
                            if job_id not in spool_copies:
                                for _ in range(10):
                                    time.sleep(0.05)
                                    if job_id in spool_copies:
                                        break
                            if job_id in spool_copies:
                                temp, _ = spool_copies.pop(job_id)
                                try:
                                    if os.path.isfile(temp):
                                        # Renomear SPL temporário para nome final
                                        os.rename(temp, dest_path)
                                        local_arquivo = dest_path
                                        # Renomear SHD companion se existir
                                        temp_shd = os.path.splitext(temp)[0] + ".shd"
                                        dest_shd = os.path.splitext(dest_path)[0] + ".shd"
                                        if os.path.isfile(temp_shd):
                                            try:
                                                os.rename(temp_shd, dest_shd)
                                            except OSError:
                                                pass
                                except OSError:
                                    if os.path.isfile(temp):
                                        local_arquivo = temp  # fallback: registra o _temp_*.spl
                            if not local_arquivo:  # fallback: cópia direta com retries
                                if copiar_spool_para_arquivos(job_id, dest_path):
                                    local_arquivo = dest_path
                                    copiar_shd_para_arquivos(job_id, dest_path)
                            if not local_arquivo:
                                log_aviso(
                                    "monitorar_impressoes.copia_spool",
                                    f"Cópia do spool não salva: job_id={job_id} documento={documento!r} impressora={printer_name!r}",
                                )
                                print(f"      [Aviso] Cópia do spool não salva (arquivo não encontrado ou sem permissão em System32\\spool\\PRINTERS)")

                        # Converter EMF → BMP ANTES de salvar no Excel para registrar caminho correto
                        imagens = []
                        if CONVERTER_EMF_PARA_IMAGEM and local_arquivo and os.path.isfile(local_arquivo):
                            imagens = converter_spl_para_imagens(local_arquivo)

                        # --- SALVAR NO EXCEL ---
                        arquivo_hoje = caminho_log_do_dia()
                        linha = [
                            job_id, usuario, data_hora, documento, paginas, printer_name,
                            _bytes_para_kb_mb_gb(tamanho),
                            imagens[0] if imagens else local_arquivo,
                        ]
                        try:
                            if not os.path.exists(arquivo_hoje):
                                wb = Workbook()
                                ws = wb.active
                                ws.title = NOME_ABA
                                ws.append(CABECALHO)
                            else:
                                wb = load_workbook(arquivo_hoje)
                                if NOME_ABA in wb.sheetnames:
                                    ws = wb[NOME_ABA]
                                else:
                                    ws = wb.create_sheet(NOME_ABA)
                                    ws.append(CABECALHO)
                            ws.append(linha)
                            wb.save(arquivo_hoje)

                            print(f"[NOVO] {data_hora} | {usuario} | {documento} ({paginas} pgs)")
                            if SALVAR_COPIA_SPL and os.path.isfile(local_arquivo):
                                print(f"      Cópia spool salva: {local_arquivo}")
                            if imagens:
                                print(f"      [EMF] {len(imagens)} imagem(ns) gerada(s) em arquivos/")

                            jobs_processados[job_unique_key] = time.time()

                        except PermissionError as e_perm:
                            log_erro("monitorar_impressoes.excel", f"Arquivo Excel aberto por outro programa: {arquivo_hoje}", e_perm)
                            print("Erro: Arquivo Excel aberto por outro programa. Feche o arquivo; a linha será gravada na próxima tentativa.")
                            # Marcar como processado AGORA para evitar duplicata: sem isso o job
                            # seria reprocessado no próximo ciclo enquanto ainda na fila, gerando
                            # duas linhas no Excel quando o arquivo fosse desbloqueado.
                            jobs_processados[job_unique_key] = time.time()
                            excel_pendente.append((linha, job_unique_key, time.time()))

                        except Exception as e_excel:
                            log_erro("monitorar_impressoes.excel", f"Falha ao salvar/abrir Excel: {arquivo_hoje}", e_excel)
                            print(f"Erro ao salvar Excel ({type(e_excel).__name__}): {e_excel}")

                except OSError as e:
                    log_aviso("monitorar_impressoes.impressora", f"Impressora inacessível: {printer_name!r} | {e}")
                    print(f"[Aviso] Impressora '{printer_name}': {e}")
                finally:
                    if p_handle:
                        win32print.ClosePrinter(p_handle)

            if DEBUG and total_jobs_ciclo > 0:
                print(f"[Diagnóstico] Total de jobs na fila neste ciclo: {total_jobs_ciclo}")

            # Limpeza de memória: remove jobs do cache com mais de 24h
            agora = time.time()
            chaves_para_remover = [k for k, v in jobs_processados.items() if agora - v > CACHE_JOB_HORAS * 3600]
            for k in chaves_para_remover:
                del jobs_processados[k]

            # Limpeza de excel_pendente muito antigos (> 24h)
            excel_pendente = [(l, c, ts) for l, c, ts in excel_pendente if agora - ts <= CACHE_JOB_HORAS * 3600]

            # Limpeza de spool_copies órfãos (watcher capturou mas loop nunca processou)
            chaves_spool_velhas = [
                k for k, (caminho, ts) in list(spool_copies.items())
                if agora - ts > SPOOL_COPIES_MAX_IDADE_SEGUNDOS
            ]
            for k in chaves_spool_velhas:
                item = spool_copies.pop(k, None)  # None se a thread watcher removeu antes
                if item is None:
                    continue
                caminho_velho, _ = item
                for ext_path in (caminho_velho, os.path.splitext(caminho_velho)[0] + ".shd"):
                    if os.path.isfile(ext_path):
                        try:
                            os.remove(ext_path)
                        except OSError:
                            pass
                log_aviso("cleanup_spool_copies", f"Temp SPL órfão removido: {caminho_velho}")

            # Limpeza de logs antigos: uma vez por dia
            if datetime.now() - ultima_limpeza > timedelta(days=1):
                limpar_logs_antigos()
                ultima_limpeza = datetime.now()

        except KeyboardInterrupt:
            raise  # Ctrl+C — encerra normalmente
        except Exception as e:
            log_erro("monitorar_impressoes.loop", "Erro no loop de monitoramento", e)
            print(f"Erro no loop de monitoramento: {type(e).__name__}: {e}")

        time.sleep(INTERVALO_SEGUNDOS)


if __name__ == "__main__":
    def _excepthook(tipo, valor, tb):
        log = configurar_log_erro()
        log.critical(
            "[main] Exceção não tratada: %s: %s\n%s",
            tipo.__name__, valor, "".join(traceback.format_exception(tipo, valor, tb)),
        )
        sys.__excepthook__(tipo, valor, tb)

    sys.excepthook = _excepthook
    try:
        iniciar_log()
        monitorar_impressoes()
    except KeyboardInterrupt:
        print("\nMonitoramento encerrado pelo usuário.")
    except Exception as e:
        log_erro("main", "Encerramento por exceção", e)
        raise
