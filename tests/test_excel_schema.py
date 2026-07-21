"""Integração leve: schema Excel sem hardware de impressão."""
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook

import app


def test_grava_linha_com_cabecalho_dec001(tmp_path, monkeypatch):
    monkeypatch.setattr(app, "PASTA_SCRIPT", str(tmp_path))
    arquivo = app.caminho_log_do_dia()
    linha = [
        1,
        "usuario_teste",
        "2026-07-21 10:00:00",
        "doc.txt",
        1,
        "Microsoft Print to PDF",
        4096,
        "4.00 KB",
        str(tmp_path / "arquivos" / "1_imp_doc.spl"),
        str(tmp_path / "arquivos" / "1_imp_doc_p001.bmp"),
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = app.NOME_ABA
    ws.append(app.CABECALHO)
    ws.append(linha)
    wb.save(arquivo)

    wb2 = load_workbook(arquivo)
    ws2 = wb2[app.NOME_ABA]
    assert list(ws2.iter_rows(min_row=1, max_row=1, values_only=True))[0] == tuple(app.CABECALHO)
    row2 = list(ws2.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    assert row2[0] == 1
    assert row2[6] == 4096
    assert row2[7] == "4.00 KB"
    assert str(row2[8]).endswith(".spl")
    assert str(row2[9]).endswith(".bmp")
